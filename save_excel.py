import pandas as pd
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
import os


def get_last_row(excel_file_path: str, sheet_name: str = "faq_history") -> int:
    """
    指定されたExcelファイルとシートにおける、追記用の次の行番号を取得する関数。

    Args:
        excel_file_path (str): Excelファイルのパス。
        sheet_name (str): シート名（デフォルトは "Sheet1"）。

    Returns:
        int: 追記すべき行番号（ヘッダーの次から数えて末尾＋1）。
    """
    if os.path.exists(excel_file_path):
        try:
            existing_df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
            startrow = len(existing_df) + 1
            print(f"既存の行数は {len(existing_df)} 行、次の書き込み行は {startrow} 行目です。")
            return startrow
        except Exception as e:
            print(f"読み込み中にエラー: {e}")
            return 0
    else:
        print("ファイルが存在しません。新規作成します。")
        return 0
    


def save_record_to_excel(data: dict):
    output_directory = r"C:\Users\nepia\OneDrive\デスクトップ\Slack_Bolt\excel"
    excel_file_name = "faq_history.xlsx"
    excel_file_path = os.path.join(output_directory, excel_file_name)
    sheet_name = "faq_history"

    df_new = pd.DataFrame([data])

    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    if os.path.exists(excel_file_path):
        with pd.ExcelWriter(
            excel_file_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="overlay"
        ) as writer:
            # 既存シートの末尾に追記するために startrow を調整
            book = load_workbook(excel_file_path)
            if sheet_name in book.sheetnames:
                startrow = get_last_row(excel_file_path, sheet_name)
                
            else:
                startrow = 0

            df_new.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=False if startrow > 0 else True,
                startrow=startrow
            )
    else:
        # 新規ファイル作成時はそのまま保存
        df_new.to_excel(excel_file_path, sheet_name=sheet_name, index=False)

    print(f"{excel_file_path} に保存されました。")
    
 

def find_column_by_header_name(sheet: Worksheet, column_name: str) -> int:
    for col in range(1, sheet.max_column + 1):
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value == column_name:
            return col
    raise ValueError(f"'{column_name}' という列が見つかりません。")


   
def find_row_by_thread_ts(sheet: Worksheet, data: dict) -> int:
    thread_ts = data.get("thread_ts")
    thread_ts_col = find_column_by_header_name(sheet, "thread_ts")
    
    for row in range(2, sheet.max_row + 1):  # 1行目はヘッダ
        cell_value = sheet.cell(row=row, column=thread_ts_col).value
        if str(cell_value) == thread_ts:
            return row
    raise ValueError("対象の thread_ts が見つかりません。")



def update_excel_row(file_path: str, sheet_name: str, data: dict):
    wb = load_workbook(file_path)
    sheet = wb[sheet_name]

    target_row = find_row_by_thread_ts(sheet, data)
    
    # ここで必要なカラムを更新（例: "selected_item" と "user_input" を更新）
    for key in ["selected_item", "user_input"]:
        col = find_column_by_header_name(sheet, key)
        sheet.cell(row=target_row, column=col).value = data[key]
    
    wb.save(file_path)
    print(f"{target_row}行目を更新しました。")

