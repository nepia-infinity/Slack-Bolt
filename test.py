import pandas as pd
from openpyxl import load_workbook
from time_utils import convert_thread_ts_to_unix
from save_excel import update_excel_row
import os


def test_save_excel():
    data = {
        "thread_ts": "1744506064.018759",
        "channel_id": "C02B2S137FC",
        "user_id": "UQEHCN01E",
        "user_query": "教えてGemini　無色転生について教えて",
        "gemini_response": "<@UQEHCN01E> さん！お疲れ様です。\n*受付日時: 2025/04/13（日）10:01 * \n\n\n無職転生 ですね。\n\n*無職転生* は 、 理不尽な死を遂げた主人公が 、 剣と魔法の異世界で生まれ変わり  、 成長していく物語です 。 人生を後悔している主人公が 、 前世の記憶と経験を活かし 、 新たな世界で様々な困難に立ち向かい 、 生き抜く姿が描かれています 。\n\n詳細はこちらをご覧ください 。\n\n<https://mushokutensei.jp/>\n"
    }

    # 保存先を指定して保存できるが、同名のファイルがあると上書きされる
    output_directory = r"C:\Users\nepia\OneDrive\デスクトップ\Slack_Bolt\excel"
    excel_file_name = "feedback_data.xlsx"
    excel_file_path = os.path.join(output_directory, excel_file_name)

    df = pd.DataFrame([data])
    df.to_excel(excel_file_path, index=False)

    print(f"データが {excel_file_path} に保存されました。")
    


def test_get_excel_output_directory():
    """Excelファイル出力先のディレクトリパスを取得する関数"""
    # ここにExcelファイルを保存したいデフォルトのディレクトリパスを設定してください
    default_excel_directory = r"C:\Users\nepia\OneDrive\デスクトップ\Slack_Bolt\excel"
    return default_excel_directory



def test_get_last_row(excel_file_path: str, sheet_name: str = "faq_history") -> int:
    """
    指定されたExcelファイルとシートにおける、追記用の次の行番号を取得する関数。

    Args:
        excel_file_path (str): Excelファイルのパス。
        sheet_name (str): シート名（デフォルトは "Sheet1"）。

    Returns:
        int: 追記すべき行番号（ヘッダーの次から数えて末尾＋1）。
    """
    if os.path.exists(excel_file_path):
        try:
            existing_df = pd.read_excel(excel_file_path, sheet_name=sheet_name)
            startrow = len(existing_df) + 1
            print(f"既存の行数は {len(existing_df)} 行、次の書き込み行は {startrow} 行目です。")
            return startrow
        except Exception as e:
            print(f"読み込み中にエラー: {e}")
            return 0
    else:
        print("ファイルが存在しません。新規作成します。")
        return 0
    


def test_save_record_to_excel(data: dict):
    output_directory = test_get_excel_output_directory()
    excel_file_name = "faq_history.xlsx"
    excel_file_path = os.path.join(output_directory, excel_file_name)
    sheet_name = "faq_history"

    df_new = pd.DataFrame([data])

    if not os.path.exists(output_directory):
        os.makedirs(output_directory)

    if os.path.exists(excel_file_path):
        with pd.ExcelWriter(
            excel_file_path,
            engine="openpyxl",
            mode="a",
            if_sheet_exists="overlay"
        ) as writer:
            # 既存シートの末尾に追記するために startrow を調整
            book = load_workbook(excel_file_path)
            if sheet_name in book.sheetnames:
                startrow = test_get_last_row(excel_file_path, sheet_name)
            else:
                startrow = 0

            df_new.to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
                header=False if startrow > 0 else True,
                startrow=startrow
            )
    else:
        # 新規ファイル作成時はそのまま保存
        df_new.to_excel(excel_file_path, sheet_name=sheet_name, index=False)

    print(f"{excel_file_path} に保存されました。")


    
def test_convert_thread_ts_to_unix():
    thread_ts = "1744593029.666519"
    result = convert_thread_ts_to_unix(thread_ts)
    


def test_update_excel_row():
    file_path = r"C:\Users\nepia\OneDrive\デスクトップ\Slack_Bolt\excel\faq_history.xlsx"
    sheet_name = "faq_history"
    data = {
        "thread_ts": "1744598437.208029",
        "formatted_ts": "2025/04/14 11:40:37",
        "channel_id": "C02B2S137FC",
        "user_id": "UQEHCN01E",
        "selected_item": "情報が古い・更新されていない",
        "user_input": "本当なわけあるかぁー"
    }

    update_excel_row(file_path, sheet_name, data)
    
    return

if __name__ == "__main__":
    test_update_excel_row()
    
    # test_convert_thread_ts_to_unix()
    
    # data = {
    #     "thread_ts": "1744506064.018759",
    #     "channel_id": "C02B2S137FC",
    #     "user_id": "UQEHCN01E",
    #     "user_query": "教えてGemini　無色転生について教えて",
    #     "gemini_response": "<@UQEHCN01E> さん！お疲れ様です。\n*受付日時: 2025/04/13（日）10:01 * \n\n\n無職転生 ですね。\n\n*無職転生* は 、 理不尽な死を遂げた主人公が 、 剣と魔法の異世界で生まれ変わり  、 成長していく物語です 。 人生を後悔している主人公が 、 前世の記憶と経験を活かし 、 新たな世界で様々な困難に立ち向かい 、 生き抜く姿が描かれています 。\n\n詳細はこちらをご覧ください 。\n\n<https://mushokutensei.jp/>\n",
    #     "is_useful": False,
    #     "selected_item": "共有されたリンクにアクセスできなかった",
    #     "expected_response" : "正しいリンクが表示されていえば完璧だった。"
    # }
    
    # test_save_record_to_excel(data)